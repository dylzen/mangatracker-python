"""One-time migration script to import source URLs from the Excel file into SQLite."""

import os
from openpyxl import load_workbook
import db

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "Manga_Collection.xlsx")

SOURCE_A_URL_COL = 26
SOURCE_B_URL_COL = 25


def migrate():
    db.init_db()

    book = load_workbook(EXCEL_PATH, read_only=True)
    sheet = book["auto"]

    source_a_urls = []
    source_b_urls = []

    for row in sheet.iter_rows(min_row=2, min_col=SOURCE_A_URL_COL, max_col=SOURCE_A_URL_COL):
        value = row[0].value
        if value:
            source_a_urls.append(value)

    for row in sheet.iter_rows(min_row=2, min_col=SOURCE_B_URL_COL, max_col=SOURCE_B_URL_COL):
        value = row[0].value
        if value:
            source_b_urls.append(value)

    book.close()

    for i, source_a_url in enumerate(source_a_urls):
        source_b_url = source_b_urls[i] if i < len(source_b_urls) else None
        db.insert_manga(source_a_url=source_a_url, source_b_url=source_b_url)
        print(f"Imported: {source_a_url}")

    for i in range(len(source_a_urls), len(source_b_urls)):
        db.insert_manga(source_b_url=source_b_urls[i])
        print(f"Imported (Source B only): {source_b_urls[i]}")

    print(f"\nMigration complete. Imported {max(len(source_a_urls), len(source_b_urls))} manga entries.")


if __name__ == "__main__":
    migrate()
